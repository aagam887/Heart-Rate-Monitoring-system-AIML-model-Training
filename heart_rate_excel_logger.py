# heart_rate_excel_logger.py
# ─────────────────────────────────────────────────────────────
# Reads CSV lines from ESP32 over Serial and saves them to
# an Excel file (.xlsx) at the configured save path.
#
# Rules:
#   • Same .xlsx file is reused every time (never overwritten)
#   • Each new session adds a 3-row gap then a date/time heading
#   • Columns: Timestamp | Session Sec | Activity | HR (BPM) |
#              HRV (ms) | MSE | Threshold | Status
#
# Install dependencies:
#   pip install pyserial openpyxl
# ─────────────────────────────────────────────────────────────

import serial
import serial.tools.list_ports
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment,
                               Border, Side)
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
import time

# ══════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════
SAVE_FOLDER  = r"D:\CTAIML FINAL WORK\HeartRateAI"
EXCEL_FILE   = os.path.join(SAVE_FOLDER, "heartrate_log.xlsx")
BAUD_RATE    = 115200

# Column definitions — order must match CSV output from ESP32
COL_HEADERS = [
    "Timestamp (ms)",
    "Session (sec)",
    "Activity",
    "HR (BPM)",
    "HRV (ms)",
    "MSE",
    "Threshold",
    "Status"
]

# Column widths in Excel units
COL_WIDTHS = [16, 14, 10, 10, 10, 10, 12, 12]

# Style constants
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
SESSION_FILL  = PatternFill("solid", fgColor="D6E4F0")
SESSION_FONT  = Font(bold=True, color="1F4E79", size=11, italic=True)
NORMAL_FONT   = Font(size=10)
ANOMALY_FILL  = PatternFill("solid", fgColor="FFE5E5")
ANOMALY_FONT  = Font(size=10, color="C00000")
CENTER_ALIGN  = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN    = Alignment(horizontal="left",   vertical="center")

thin = Side(style="thin", color="AAAAAA")
CELL_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


# ══════════════════════════════════════════════════════════════
# EXCEL HELPERS
# ══════════════════════════════════════════════════════════════

def get_or_create_workbook():
    """Load existing workbook or create a new one with headers."""
    os.makedirs(SAVE_FOLDER, exist_ok=True)

    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        print(f"Loaded existing file: {EXCEL_FILE}")
        print(f"  Current rows: {ws.max_row}")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Heart Rate Log"

        # Write column headers on row 1
        for col_idx, (header, width) in enumerate(
                zip(COL_HEADERS, COL_WIDTHS), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border    = CELL_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"   # freeze header row

        wb.save(EXCEL_FILE)
        print(f"Created new file: {EXCEL_FILE}")

    return wb, ws


def add_session_heading(ws, session_label):
    """
    Adds 3 blank rows then a merged session heading row.
    Returns the row number of the heading.
    """
    next_row = ws.max_row + 1

    # 3 blank gap rows
    for _ in range(3):
        ws.cell(row=next_row, column=1, value="")
        next_row += 1

    # Session heading — merged across all columns
    n_cols = len(COL_HEADERS)
    ws.merge_cells(
        start_row=next_row, start_column=1,
        end_row=next_row, end_column=n_cols
    )
    cell = ws.cell(row=next_row, column=1, value=session_label)
    cell.font      = SESSION_FONT
    cell.fill      = SESSION_FILL
    cell.alignment = CENTER_ALIGN
    cell.border    = CELL_BORDER
    ws.row_dimensions[next_row].height = 18

    return next_row + 1   # first data row after heading


def write_data_row(ws, row_num, values, is_anomaly=False):
    """Write one data row with styling."""
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.border    = CELL_BORDER
        cell.alignment = CENTER_ALIGN if col_idx != 3 else LEFT_ALIGN

        if is_anomaly:
            cell.font = ANOMALY_FONT
            cell.fill = ANOMALY_FILL
        else:
            cell.font = NORMAL_FONT

    ws.row_dimensions[row_num].height = 16


def save_workbook(wb):
    try:
        wb.save(EXCEL_FILE)
    except PermissionError:
        backup = EXCEL_FILE.replace(".xlsx", f"_backup_{int(time.time())}.xlsx")
        wb.save(backup)
        print(f"File locked — saved backup to: {backup}")


# ══════════════════════════════════════════════════════════════
# SERIAL PORT DETECTION
# ══════════════════════════════════════════════════════════════

def find_esp32_port():
    ports = serial.tools.list_ports.comports()
    for port in ports:
        desc = port.description.upper()
        if any(k in desc for k in
               ["CH340", "CP210", "SILABS", "USB SERIAL", "UART"]):
            return port.device
    print("Available ports:")
    for p in ports:
        print(f"  {p.device}  —  {p.description}")
    return None


# ══════════════════════════════════════════════════════════════
# MAIN LOGGER
# ══════════════════════════════════════════════════════════════

def main():
    print("=" * 55)
    print("Heart Rate AI — Excel Logger")
    print("=" * 55)

    # Find port
    port = find_esp32_port()
    if not port:
        port = input(
            "\nCould not auto-detect port. Enter COM port "
            "(e.g. COM4): "
        ).strip()

    print(f"\nConnecting to {port} at {BAUD_RATE} baud...")
    try:
        ser = serial.Serial(port, BAUD_RATE, timeout=2)
        time.sleep(2)
        print("Connected")
    except Exception as e:
        print(f"Connection failed: {e}")
        return

    # Load or create workbook
    wb, ws = get_or_create_workbook()

    # Add session heading
    now = datetime.now()
    session_label = (
        f"Session: {now.strftime('%d-%b-%Y  %I:%M %p')}  |  "
        f"File: {EXCEL_FILE}"
    )
    next_data_row = add_session_heading(ws, session_label)
    save_workbook(wb)

    print(f"\nSaving to  : {EXCEL_FILE}")
    print(f"Session row: {next_data_row - 1}")
    print("Press Ctrl+C to stop\n")

    # Print console header
    print(f"{'Time':>20} {'Act':>6} {'HR':>6} "
          f"{'HRV':>6} {'MSE':>8} {'Status':>10}")
    print("-" * 65)

    rows_written = 0
    anomaly_count = 0
    current_row = next_data_row

    # Save every N rows to avoid data loss on crash
    SAVE_EVERY = 5

    try:
        while True:
            try:
                raw = ser.readline()
                if not raw:
                    continue
                line = raw.decode("utf-8", errors="ignore").strip()
            except Exception:
                continue

            if not line:
                continue

            # Print non-CSV lines (debug output) without processing
            if not line.startswith("CSV,"):
                print(f"  ESP32: {line}")
                continue

            # Parse CSV line
            # Format: CSV,<ms>,<sec>,<act>,<hr>,<hrv>,<mse>,<thresh>,<status>
            parts = line.split(",")
            if len(parts) != 9:
                continue

            try:
                _, ts_ms, sess_s, activity, hr, hrv, mse, thresh, status = parts
                ts_ms  = int(ts_ms)
                sess_s = int(sess_s)
                hr_val = float(hr)
                hrv_val = float(hrv)
                mse_val = float(mse)
                thresh_val = float(thresh)
                status = status.strip()
            except (ValueError, IndexError):
                continue

            is_anomaly = (status != "Normal")
            if is_anomaly:
                anomaly_count += 1

            # Values in column order matching COL_HEADERS
            values = [
                ts_ms,
                sess_s,
                activity,
                round(hr_val, 1),
                round(hrv_val, 1),
                round(mse_val, 4),
                round(thresh_val, 4),
                status
            ]

            write_data_row(ws, current_row, values, is_anomaly)
            current_row += 1
            rows_written += 1

            # Console output
            now_str = datetime.now().strftime("%H:%M:%S")
            anomaly_flag = " *** ANOMALY ***" if is_anomaly else ""
            print(f"  {now_str:>10} {activity:>6} "
                  f"{hr_val:>6.1f} {hrv_val:>6.1f} "
                  f"{mse_val:>8.4f} {status:>10}"
                  f"{anomaly_flag}")

            # Periodic save
            if rows_written % SAVE_EVERY == 0:
                save_workbook(wb)

    except KeyboardInterrupt:
        pass

    # Final save
    save_workbook(wb)
    ser.close()

    print(f"\n{'='*55}")
    print(f"Session ended")
    print(f"  Rows written : {rows_written}")
    print(f"  Anomalies    : {anomaly_count}")
    print(f"  File         : {EXCEL_FILE}")
    print(f"{'='*55}")


if __name__ == "__main__":
    main()
